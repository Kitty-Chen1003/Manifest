import re
import shutil
from dataclasses import dataclass
from typing import Optional, Callable

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill


@dataclass(frozen=True)
class FieldRule:
    max_length: Optional[int] = None
    pattern: Optional[str] = None
    custom_check: Optional[Callable] = None


# UNICODE_TEXT = r"^[\w\d\s]+$"
# UNICODE_TEXT = r"^[\p{L}\p{N}\p{P}\s]+$"
NUMERIC = r"^\d+$"

FIELD_RULES = {

    "InvoiceNumber": FieldRule(
        max_length=70
    ),

    "GoodsItemNumber": FieldRule(
        max_length=5,
        pattern=NUMERIC
    ),

    "HSCode": FieldRule(
        pattern=r"^[\w\d]{6,}$"
    ),

    "Total Price": FieldRule(
        max_length=17,
        pattern=r"^\d+(\.\d{1,2})?$"
    ),

    "GrossMassKg": FieldRule(
        max_length=17,
        pattern=r"^\d+(\.\d{1,6})?$"
    ),

    "CountryOriginCode": FieldRule(
        pattern=r"^.{2}$"
    ),

    "AmountPackages": FieldRule(
        max_length=8,
        pattern=NUMERIC
    ),

    "TrackingNumber": FieldRule(
        max_length=35
    ),

    "InvoiceCurrency": FieldRule(
        pattern=r"^.{3}$"
    ),

    "DescriptionGoods": FieldRule(
        max_length=512
    ),

    "ConsignorName": FieldRule(
        max_length=70
    ),

    "ConsignorStreetAndNr": FieldRule(
        max_length=70
    ),

    "ConsignorCity": FieldRule(
        max_length=35
    ),

    "ConsignorPostcode": FieldRule(
        max_length=17
    ),

    "ConsignorCountry": FieldRule(
        pattern=r"^.{2}$"
    ),

    "ConsigneeName": FieldRule(
        max_length=70
    ),

    "ConsigneeStreetAndNr": FieldRule(
        max_length=70
    ),

    "ConsigneePostcode": FieldRule(
        max_length=17,
        custom_check=lambda df: ~(
                (df["ConsigneeCountryCode"].str.upper() == "PL") &
                (~df["ConsigneePostcode"].str.match(r"^.{2}-.{3}$", na=False))
        )
    ),

    "ConsigneeCity": FieldRule(
        max_length=35
    ),

    "ConsigneeCountryCode": FieldRule(
        pattern=r"^.{2}$"
    ),

    "INCOTerm": FieldRule(
        pattern=r"^.{3}$"
    ),

    "AirWayBill": FieldRule(
        max_length=70
    ),

    "IOSS": FieldRule(
        max_length=17
    ),

    "Box Number": FieldRule(
        max_length=70
    ),
}


def handle_pl_postcode(df):
    # 保证 _has_error 存在
    if "_has_error" not in df.columns:
        df["_has_error"] = False

    if "_is_fixed" not in df.columns:
        df["_is_fixed"] = False

    country = df["ConsigneeCountryCode"].astype(str).str.upper()
    mask_pl = country == "PL"

    # 邮编仅用于判断（不污染原数据）
    postcode_check = df["ConsigneePostcode"].astype(str)

    valid = postcode_check.str.match(r"^\d{2}-\d{3}$", na=False)
    five_digits = postcode_check.str.match(r"^\d{5}$", na=False)

    # 自动修正 12345 → 12-345
    fix_mask = mask_pl & (~valid) & five_digits

    df.loc[fix_mask, "_is_fixed"] = True

    df.loc[fix_mask, "ConsigneePostcode"] = (
            postcode_check.loc[fix_mask].str[:2]
            + "-"
            + postcode_check.loc[fix_mask].str[2:]
    )


def truncate_min_length_field(df, field, length):
    if field not in df.columns:
        return

    series = df[field].astype(str).str.strip()
    mask_long = series.str.len() > length

    df.loc[mask_long, "_is_fixed"] = True

    df.loc[mask_long, field] = series.str[:length]


def validate_dataframe(df):
    errors = []

    for field, rule in FIELD_RULES.items():
        if field not in df.columns:
            continue

        series = df[field].astype(str).str.strip()

        # 最大长度
        if rule.max_length:
            invalid_idx = series[series.str.len() > rule.max_length].index.tolist()
            if invalid_idx:
                errors.append(f"{field} exceeds max length {rule.max_length} at rows {excel_rows(invalid_idx)}")
                df.loc[invalid_idx, "_has_error"] = True

        # 正则格式（使用 re 模块）
        if rule.pattern:
            pattern = re.compile(rule.pattern)
            invalid_idx = series[series != ''].index[
                series[series != ''].apply(lambda x: not bool(pattern.fullmatch(x)))].tolist()
            if invalid_idx:
                errors.append(f"{field} format invalid at rows {excel_rows(invalid_idx)}")
                df.loc[invalid_idx, "_has_error"] = True

        # 自定义规则
        if rule.custom_check:
            try:
                result = rule.custom_check(df)

                if isinstance(result, pd.Series):
                    invalid_idx = result[result == False].index.tolist()
                    if invalid_idx:
                        errors.append(f"{field} custom rule failed at rows {excel_rows(invalid_idx)}")
                        df.loc[invalid_idx, "_has_error"] = True

                elif result is False:
                    errors.append(f"{field} custom rule failed")
                    df["_has_error"] = True

            except Exception as e:
                print(e)
                errors.append(f"{field} custom rule execution error")
                df["_has_error"] = True

    if "_is_fixed" in df.columns:
        fixed_idx = df[df["_is_fixed"]].index.tolist()
        if fixed_idx:
            errors.append(f"Data auto-fixed at rows {excel_rows(fixed_idx)}")

    return errors


def excel_rows(idx_list):
    return [i + 2 for i in idx_list]


def export_error_excel(original_excel_path, df, save_path):
    shutil.copy(original_excel_path, save_path)

    wb = load_workbook(save_path)
    ws1 = wb.active
    ws1.title = "Sheet1"

    red_fill = PatternFill(
        start_color="FFFFC7CE",
        end_color="FFFFC7CE",
        fill_type="solid"
    )

    yellow_fill = PatternFill(
        start_color="FFFFEB9C",
        end_color="FFFFEB9C",
        fill_type="solid"
    )

    for df_idx in df.index:
        excel_row = df_idx + 2

        if df.loc[df_idx, "_has_error"]:
            fill = red_fill
        elif "_is_fixed" in df.columns and df.loc[df_idx, "_is_fixed"]:
            fill = yellow_fill
        else:
            continue

        for col in range(1, ws1.max_column + 1):
            ws1.cell(row=excel_row, column=col).fill = fill

    wb.save(save_path)
