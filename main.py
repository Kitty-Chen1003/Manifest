import os
import sys
import time
from datetime import datetime, timedelta

import pytz
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, \
    QHeaderView, QAbstractItemView, QMessageBox, QDialog, QAction, QToolBar, QHBoxLayout, QLabel, \
    QScrollArea, QComboBox, QLineEdit, QPushButton, QMenu
from PyQt5.QtCore import Qt
import pandas as pd

from PyQt5.QtGui import QPixmap, QImage

import fitz  # PyMuPDF
from openpyxl.styles import Font
from openpyxl.workbook import Workbook

from utils import pdf, db
from utils.common import normalize

from views.create_sads import CreateSADs
from views.sad import SADWindow
from views.import_dictionary import DictionarySelectorDialog

import json

from views.send_zc415 import SendZC415
from views.send_message_by_main_id import SendMessageByMainID
from views.generate_pdf_and_xml import GeneratePDFAndXML
from views.change_cl380_by_main_id import ChangeCL380ByMainID
from views.refresh_by_main_id import RefreshByMainId
# from views.clear_data import ClearData
from views.clear_data_by_time import DateTimeDialog

from communication import http_client
from views.login import LoginDialog

from utils import manifest


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.current_sub_df = None
        self.username = None
        self.token = None

        # 设置窗口标题和大小
        self.setWindowTitle('Customs Declaration System')
        self.resize(1260, 1000)

        # 0代表main_table，1代表sub_table
        self.index_table = 0

        # self.list_data_rows = []

        self.list_fixed_key = ['InvoiceNumber', 'GoodsItemNumber', 'HSCode', 'Total Price', 'GrossMassKg',
                               'AmountPackages', 'TrackingNumber', 'ConsignorName', 'InvoiceCurrency',
                               'DescriptionGoods', 'ConsignorStreetAndNr', 'ConsignorCity', 'ConsignorPostcode',
                               'ConsignorCountry', 'ConsigneeName', 'ConsigneeStreetAndNr', 'ConsigneePostcode',
                               'ConsigneeCity', 'ConsigneeCountryCode', 'AirWayBill', 'IOSS', 'CountryOriginCode',
                               'ConsigneeNameID', 'Box Number', 'Dsk']

        db.create_tables()

        self.tool_button = None
        self.list_time_out_ids = {}

        self.flag = 0

        self.datetime = None
        self.time_flag = 0

        self.input_information = None

        self.main_table_id = 0
        self.sub_table_id = 0
        # self.data_rows = []

        # 初始化 DataFrame
        self.df_main = pd.DataFrame({
            'AirWayBill': pd.Series(dtype='str'),
            'created date': pd.Series(dtype='str'),
            'state': pd.Series(dtype='str')
        })

        self.df_sub = pd.DataFrame({
            'IOSS': pd.Series(dtype='str'),
            'TrackingNumber': pd.Series(dtype='str'),
            'lrn': pd.Series(dtype='str'),
            'state': pd.Series(dtype='str'),
            'event time': pd.Series(dtype='str')
        })

        self.df_response = pd.DataFrame({
            'ID': pd.Series(dtype='str'),
            'type': pd.Series(dtype='str'),
            'message_id': pd.Series(dtype='str'),
            'event time': pd.Series(dtype='str'),
            'direction': pd.Series(dtype='str')
        })

        self.df_upd = pd.DataFrame({
            'ID': pd.Series(dtype='str'),
            'type': pd.Series(dtype='str'),
            'message_id': pd.Series(dtype='str'),
            'event time': pd.Series(dtype='str'),
            'direction': pd.Series(dtype='str')
        })

        self.detail_response_data = []
        self.detail_upd_data = []

        self.df_main_data_list = []
        self.df_sub_data_list = []
        self.df_upd_data_list = []

        # 创建滚动区域
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)  # 允许自动调整内容大小

        # 创建中央控件
        self.central_widget = QWidget()
        self.scroll_area.setWidget(self.central_widget)  # 将 central_widget 设置为滚动区域的内容

        # 设置滚动区域为主窗口的中央控件
        self.setCentralWidget(self.scroll_area)

        # 创建布局
        self.layout = QVBoxLayout(self.central_widget)

        # 创建菜单栏
        # self.create_menubar()
        # 创建工具栏

        self.create_toolbar()

        self.top = QWidget()
        self.top.setFixedWidth(1260)
        self.top.setFixedHeight(400)

        self.layout_top = QHBoxLayout(self.top)

        self.main_table = QTableWidget()
        self.main_table.setMaximumWidth(700)
        self.main_table.setFixedHeight(350)
        self.sub_table = QTableWidget()
        self.sub_table.setMaximumWidth(700)
        self.sub_table.setFixedHeight(350)

        self.layout_top.addWidget(self.main_table)
        # self.layout_top.addWidget(self.sub_table)

        self.sub_table_container = QWidget()
        self.sub_table_layout = QVBoxLayout(self.sub_table_container)
        self.sub_table_layout.setContentsMargins(0, 0, 0, 0)
        self.sub_search_widget = QWidget()
        self.sub_search_layout = QHBoxLayout(self.sub_search_widget)
        self.sub_search_layout.setContentsMargins(0, 0, 0, 0)

        self.sub_search_input = QLineEdit()
        self.sub_search_input.setFixedWidth(260)  # 设置固定宽度

        self.sub_search_button = QPushButton("Search")
        self.sub_search_layout.addWidget(self.sub_search_input)
        self.sub_search_layout.addWidget(self.sub_search_button)
        self.sub_search_layout.addStretch()  # 右侧留空

        self.sub_table_layout.addWidget(self.sub_search_widget)

        # 添加 sub_table
        self.sub_table_layout.addWidget(self.sub_table)

        # 默认隐藏整个容器
        self.sub_table_container.hide()

        # 把容器加入 top 布局（替代原来的 self.sub_table）
        self.layout_top.addWidget(self.sub_table_container)

        self.sub_search_button.clicked.connect(self.search_sub_table)
        self.sub_search_input.returnPressed.connect(self.search_sub_table)

        # ================================
        # UPD table container (new)
        # ================================
        self.upd_container = QWidget()
        self.upd_container_layout = QVBoxLayout(self.upd_container)
        self.upd_container_layout.setContentsMargins(0, 0, 0, 0)

        # ----- Filter UI -----
        self.upd_filter_widget = QWidget()
        self.upd_filter_layout = QHBoxLayout(self.upd_filter_widget)
        self.upd_filter_layout.setContentsMargins(0, 0, 0, 0)

        # Type filter
        self.upd_type_filter = QComboBox()
        self.upd_type_filter.addItem("All")
        for t in ["upd", "zc429", "zc460", "zcx03", "zcx64", "zcx65", "zc410"]:
            self.upd_type_filter.addItem(t)

        # Date filter
        self.upd_date_filter = QComboBox()
        self.upd_date_filter.addItems(["All", "Today", "Last 7 Days"])

        # Connect filters to refresh function
        self.upd_type_filter.currentIndexChanged.connect(lambda: self.update_upd_table())
        self.upd_date_filter.currentIndexChanged.connect(lambda: self.update_upd_table())

        # Add widgets to filter row
        self.upd_filter_layout.addWidget(QLabel("Type:"))
        self.upd_filter_layout.addWidget(self.upd_type_filter)
        self.upd_filter_layout.addSpacing(18)
        self.upd_filter_layout.addWidget(QLabel("Date Range:"))
        self.upd_filter_layout.addWidget(self.upd_date_filter)
        self.upd_filter_layout.addStretch()

        # Add filter row to container
        self.upd_container_layout.addWidget(self.upd_filter_widget)

        # --------------------
        # UPD table (original)
        # --------------------
        self.upd_table = QTableWidget()
        self.upd_table.setMaximumWidth(560)
        self.upd_table.setFixedHeight(350)

        self.upd_container_layout.addWidget(self.upd_table)
        self.layout_top.addWidget(self.upd_container)
        # self.upd_table = QTableWidget()
        # self.upd_table.setMaximumWidth(560)
        # self.upd_table.setFixedHeight(350)
        # self.layout_top.addWidget(self.upd_table)

        self.response = QWidget()
        self.response.setFixedWidth(1260)
        self.response.setFixedHeight(400)
        # self.response.setStyleSheet("padding: 0; border: 1px solid #222;")
        self.create_response_view()

        self.layout.addWidget(self.top)
        self.layout.addWidget(self.response)

        self.login()

        self.update_main_table()

        self.update_upd_table()

        # self.update_sub_table()

        # self.sub_table.hide()  # 初始时隐藏子表格

        self.in_bulk_delete_mode = False  # 记录是否在批量删除模式下

        # 设置列宽自动调整
        self.main_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.main_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        # 设置选择模式为单选
        self.main_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.main_table.setSelectionBehavior(QAbstractItemView.SelectRows)

        # 双击事件处理
        self.main_table.cellDoubleClicked.connect(lambda row: self.show_sub_table(row))
        self.main_table.clicked.connect(self.main_table_single_click)

        self.main_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.main_table.customContextMenuRequested.connect(self.main_table_context_menu)

        # 设置列宽自动调整
        self.sub_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.sub_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        # 设置选择模式为单选
        self.sub_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.sub_table.setSelectionBehavior(QAbstractItemView.SelectRows)

        # 双击事件处理
        self.sub_table.doubleClicked.connect(self.show_details)

        # 单击事件处理
        self.sub_table.clicked.connect(self.sub_table_single_click)

        self.sub_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.sub_table.customContextMenuRequested.connect(self.sub_table_context_menu)

        # 设置列宽自动调整
        self.upd_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.upd_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # 设置选择模式为单选
        self.upd_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.upd_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        # 单击事件处理
        self.upd_table.clicked.connect(self.upd_table_single_click)

    def create_response_view(self):
        self.layout_response = QHBoxLayout(self.response)

        self.response_table = QTableWidget()
        self.response_table.setFixedWidth(590)
        self.response_table.setFixedHeight(380)
        # 设置列宽自动调整
        self.response_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.response_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # 设置选择模式为单选
        self.response_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.response_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.layout_response.addWidget(self.response_table)
        self.update_response_table()

        # 单击事件处理
        self.response_table.clicked.connect(self.response_table_single_click)

        # 右边的可滚动自定义显示区域
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)

        # 创建一个 QWidget 作为 scroll_area 的容器
        self.info_widget = QWidget()
        # self.info_widget.setFixedWidth(590)
        self.info_layout = QVBoxLayout(self.info_widget)

        self.scroll_area.setWidget(self.info_widget)
        self.layout_response.addWidget(self.scroll_area)

    def main_table_single_click(self):
        for i in reversed(range(self.info_layout.count())):
            widget = self.info_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()  # 删除控件
                self.info_layout.removeWidget(widget)

        # 获取当前选中的行号
        row = self.main_table.currentRow()  # 或者使用 index.row()

        # 获取整行数据
        column_count = self.main_table.columnCount()
        # print(column_count)
        # 直接获取第二个单元格的数据（列索引为 0）
        if column_count > 0:  # 确保至少有一个列
            row_sequence_number = self.df_main_data_list[row][1]
            datas_cr = db.get_cr_xml_data_by_main_id(row_sequence_number, self.username)
            datas_ord = db.get_odrzucenieKomunikatu_xml_data_by_main_id(row_sequence_number, self.username)
            # print(response_data_main_id)
            self.df_response = pd.DataFrame({
                'ID': pd.Series(dtype='str'),
                'type': pd.Series(dtype='str'),
                'message_id': pd.Series(dtype='str'),
                'event time': pd.Series(dtype='str'),
                'direction': pd.Series(dtype='str')
            })
            # 遍历列表中的元组
            self.detail_response_data = []
            def process_response_datas(datas):
                for data in datas:
                    # 提取元组中的所需元素
                    response_data_id = data[0]
                    type_value = data[3]  # 元组中的第3个元素
                    replay_time = data[5]  # 元组中的第5个元素
                    direction = data[6]
                    message_id = data[10]

                    # 创建一个新的DataFrame用于存储当前行
                    new_row = pd.DataFrame({
                        'ID': [response_data_id],
                        'type': [type_value],
                        'message_id': [message_id],
                        'event time': [replay_time],
                        'direction': [direction]
                    })

                    # 使用pd.concat来合并DataFrame
                    self.df_response = pd.concat([self.df_response, new_row], ignore_index=True)

                    # 将第1和第4个元素以列表的形式存入self.detail_response_data
                    self.detail_response_data.append(
                        [data[0], data[4], data[1], data[10], data[6], data[2], data[3]])  # data[0]是第1个元素，data[3]是第4个元素

            process_response_datas(datas_cr)
            process_response_datas(datas_ord)
            self.update_response_table()


    def main_table_context_menu(self, pos):
        menu = QMenu(self.main_table)
        refresh_action = QAction("refresh UPD table", self)
        menu.addAction(refresh_action)
        refresh_action.triggered.connect(self.refresh_upd_for_selected_row)
        menu.exec_(self.main_table.viewport().mapToGlobal(pos))

    def sub_table_context_menu(self, pos):
        menu = QMenu(self.sub_table)
        get_mrn_action = QAction("query zc428-mrn", self)
        menu.addAction(get_mrn_action)
        get_mrn_action.triggered.connect(self.query_zc428_mrn_for_selected_row)
        menu.exec_(self.sub_table.viewport().mapToGlobal(pos))

    def refresh_upd_for_selected_row(self):
        row = self.main_table.currentRow()
        if row >= 0:
            main_id = self.df_main_data_list[row][1]
            current_upd_type = self.upd_type_filter.currentText()
            if current_upd_type in ["zc460", "zc429"]:
                self.update_upd_table(main_id=main_id)

    def query_zc428_mrn_for_selected_row(self):
        row = self.sub_table.currentRow()
        if row >= 0:
            lrn_number = self.current_sub_df.iloc[row]['lrn']
            if not lrn_number:
                QMessageBox.warning(
                    self,
                    "Warning",
                    "No valid LRN number found in the selected row."
                )
                return
            self.query_zc428_mrn_by_lrn(lrn_number)

    def query_zc428_mrn_by_lrn(self, lrn):
        row = db.query_zc428_mrn_by_lrn(lrn)
        if not row:
            QMessageBox.warning(
                self,
                "Warning",
                f"No zc428 data found for LRN {lrn}."
            )
            return

        xml_json_data = row[0]

        try:
            data = json.loads(xml_json_data)
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to parse JSON data: {e}"
            )
            return

        hub = data.get("p:ZC428HUB", {})
        declarations = hub.get("Declaration", [])

        for decl in declarations:
            current_lrn = decl.get("lrn", {}).get("value")
            if current_lrn == lrn:
                mrn = decl.get("mrn", {}).get("value", "")
                msg = QMessageBox(self)
                msg.setWindowTitle("Query Result")
                msg.setText(f"LRN: {lrn}\nMRN: {mrn}")
                msg.setTextInteractionFlags(Qt.TextSelectableByMouse)
                msg.exec_()
                return

        QMessageBox.information(
            self,
            "Result",
            f"No MRN corresponding to LRN {lrn} was found in zc428."
        )

    # sub_table单击事件，显示response_table
    def sub_table_single_click(self):

        for i in reversed(range(self.info_layout.count())):
            widget = self.info_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()  # 删除控件
                self.info_layout.removeWidget(widget)

        # 获取当前选中的行号
        row = self.sub_table.currentRow()  # 或者使用 index.row()

        # 获取整行数据
        column_count = self.sub_table.columnCount()
        # print(column_count)
        # 直接获取第二个单元格的数据（列索引为 0）
        if column_count > 0:  # 确保至少有一个列
            row_sequence_number = self.current_sub_df.iloc[row]['sub_id']
            response_data_sub_id = db.get_sub_xml_data_by_sub_table_id(row_sequence_number, self.username)
            # print(response_data_main_id)
            self.df_response = pd.DataFrame({
                'ID': pd.Series(dtype='str'),
                'type': pd.Series(dtype='str'),
                'message_id': pd.Series(dtype='str'),
                'event time': pd.Series(dtype='str'),
                'direction': pd.Series(dtype='str')
            })
            # 遍历列表中的元组
            self.detail_response_data = []
            for data in response_data_sub_id:
                # 提取元组中的所需元素
                response_data_id = data[0]
                type_value = data[3]  # 元组中的第3个元素
                replay_time = data[5]  # 元组中的第5个元素
                direction = data[6]
                message_id = data[10]

                # 创建一个新的DataFrame用于存储当前行
                new_row = pd.DataFrame({
                    'ID': [response_data_id],
                    'type': [type_value],
                    'message_id': [message_id],
                    'event time': [replay_time],
                    'direction': [direction]
                })

                # 使用pd.concat来合并DataFrame
                self.df_response = pd.concat([self.df_response, new_row], ignore_index=True)

                # 将第1和第4个元素以列表的形式存入self.detail_response_data
                self.detail_response_data.append(
                    [data[0], data[4], data[1], data[10], data[6], data[2], data[3]])  # data[0]是第1个元素，data[3]是第4个元素
            self.update_response_table()

    def response_table_single_click(self):
        row = self.response_table.currentRow()
        xml_data = json.loads(self.detail_response_data[row][1])
        message_id = self.detail_response_data[row][3]
        direction = self.detail_response_data[row][4]
        related_id = self.detail_response_data[row][5]
        type_value = self.detail_response_data[row][6]
        self.display_pdf(xml_data, message_id, direction, related_id, type_value)
        if type_value == "OdrzucenieKomunikatu":
            self.show_sub_table(self.main_table.currentRow(), related_id)

    def upd_table_single_click(self):
        row = self.upd_table.currentRow()
        xml_data = json.loads(self.detail_upd_data[row][1])
        message_id = self.detail_upd_data[row][3]
        direction = self.detail_upd_data[row][4]
        related_id = self.detail_upd_data[row][5]
        type_value = self.detail_upd_data[row][6]
        self.display_pdf(xml_data, message_id, direction, related_id, type_value)

    def display_pdf(self, data, message_id, direction, related_id, type_value):
        # 先清空布局中的所有控件
        for i in reversed(range(self.info_layout.count())):
            widget = self.info_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()  # 删除控件
                self.info_layout.removeWidget(widget)  # 从布局中移除控件
        # print(db.get_signature_info_by_message_id_and_username(message_id, self.username))
        signature_info = json.loads(
            db.get_signature_info_by_message_id_direction_and_username(
                message_id,
                direction,
                self.username,
                related_id,
                type_value
            )
        )

        if direction == 'send':
            flag = 0
        else:
            flag = 1
            if type_value == 'upd':
                flag = 2
        pdf_data = pdf.dict_to_pdf(data, signature_info, flag=flag)
        # 使用 PyMuPDF 打开 PDF 数据
        pdf_document = fitz.open("pdf", pdf_data)
        # 将 PDF 的每一页转换为 QPixmap
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            pix = page.get_pixmap()  # 获取页面的位图

            # 将位图转换为 QImage
            img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format_RGB888)

            # 将 QImage 转换为 QPixmap
            qpixmap = QPixmap.fromImage(img)

            label = QLabel()
            label.setPixmap(qpixmap)
            self.info_layout.addWidget(label)  # 添加到布局中

    def display_pdf_upd(self, data, signeture_data, direction):
        # 先清空布局中的所有控件
        for i in reversed(range(self.info_layout.count())):
            widget = self.info_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()  # 删除控件
                self.info_layout.removeWidget(widget)  # 从布局中移除控件
        pdf_data = pdf.dict_to_pdf(data, signeture_data)
        # 使用 PyMuPDF 打开 PDF 数据
        pdf_document = fitz.open("pdf", pdf_data)

        # 将 PDF 的每一页转换为 QPixmap
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            pix = page.get_pixmap()  # 获取页面的位图

            # 将位图转换为 QImage
            img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format_RGB888)

            # 将 QImage 转换为 QPixmap
            qpixmap = QPixmap.fromImage(img)

            label = QLabel()
            label.setPixmap(qpixmap)
            self.info_layout.addWidget(label)  # 添加到布局中

    def update_response_table(self):
        self.response_table.setRowCount(len(self.df_response))
        self.response_table.setColumnCount(len(self.df_response.columns))
        self.response_table.setHorizontalHeaderLabels(self.df_response.columns)

        for row in range(len(self.df_response)):
            for col in range(len(self.df_response.columns)):
                item_text = str(self.df_response.iloc[row, col]) if pd.notna(self.df_response.iloc[row, col]) else ''
                item = QTableWidgetItem(item_text)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # 设置为只读
                self.response_table.setItem(row, col, item)

    def update_main_table(self):
        # 初始化 DataFrame
        self.df_main = pd.DataFrame({
            'AirWayBill': pd.Series(dtype='str'),
            'created date': pd.Series(dtype='str'),
            'state': pd.Series(dtype='str')
        })
        self.df_main_data_list = db.get_main_table_data(self.username)

        # 将数据填入 DataFrame
        for data in self.df_main_data_list:
            created_date = data[2]
            state = data[3]
            AirWayBill = data[7]

            # 追加新行到 df_main
            self.df_main = pd.concat([self.df_main, pd.DataFrame({
                'AirWayBill': [AirWayBill],
                'created date': [created_date],
                'state': [state]
            })], ignore_index=True)

        self.main_table.setRowCount(len(self.df_main))
        self.main_table.setColumnCount(len(self.df_main.columns))
        self.main_table.setHorizontalHeaderLabels(self.df_main.columns)

        for row in range(len(self.df_main)):
            for col in range(len(self.df_main.columns)):
                item_text = str(self.df_main.iloc[row, col]) if pd.notna(self.df_main.iloc[row, col]) else ''
                item = QTableWidgetItem(item_text)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # 设置为只读
                self.main_table.setItem(row, col, item)

    def search_sub_table(self):
        keyword = self.sub_search_input.text().strip()
        if not keyword:
            # 搜索框为空时显示完整子表
            self.update_sub_table()
            return

        # 按行过滤，只要某一列包含关键字就保留
        mask = self.df_sub.apply(
            lambda row: row.astype(str).str.contains(keyword, case=False).any(),
            axis=1
        )
        filtered_df = self.df_sub[mask].reset_index(drop=True)

        # 更新子表显示
        self.update_sub_table(filtered_df)

    def update_sub_table(self, df=None):
        if df is None:
            df = self.df_sub

        self.current_sub_df = df if df is not None else self.df_sub

        self.sub_table.setRowCount(len(df))
        self.sub_table.setColumnCount(len(df.columns))
        self.sub_table.setHorizontalHeaderLabels(df.columns)

        self.sub_table.setUpdatesEnabled(False)
        self.sub_table.blockSignals(True)
        self.sub_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.sub_table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)

        values = df.values
        rows, cols = values.shape

        for row in range(rows):
            for col in range(cols):
                # item_text = str(df.iloc[row, col]) if pd.notna(df.iloc[row, col]) else ''
                # item = QTableWidgetItem(item_text)
                # item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # 设置为只读
                # self.sub_table.setItem(row, col, item)
                val = "" if pd.isna(values[row][col]) else str(values[row][col])
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # 设置为只读
                self.sub_table.setItem(row, col, item)

        self.sub_table.blockSignals(False)
        self.sub_table.setUpdatesEnabled(True)
        self.sub_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.sub_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

    def update_upd_table(self, main_id=None):

        # -------------------------
        # Step 0: read filter values directly from UI
        # -------------------------
        selected_type = self.upd_type_filter.currentText()
        selected_date_option = self.upd_date_filter.currentText()

        pl_tz = pytz.timezone("Europe/Warsaw")
        pl_now = datetime.now(pl_tz)

        if selected_date_option == "Today":
            date_start = pl_now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif selected_date_option == "Last 7 Days":
            date_start = (pl_now - timedelta(days=6)).replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            date_start = None  # "All" 不限制时间

        # -------------------------
        # Step 1：获取数据库数据
        # -------------------------
        all_types = ['upd', 'zc429', "zcx03", "zcx64", "zcx65", "zc410", "zc460"]
        types = [selected_type] if selected_type != "All" else all_types

        if date_start is not None:
            upd_datas = db.get_account_data(self.username, types, date_start.strftime("%Y-%m-%d %H:%M:%S"))
        elif self.time_flag:
            upd_datas = db.get_account_data(self.username, types, self.datetime)
        else:
            upd_datas = db.get_account_data(self.username, types)

        if main_id is not None:
            upd_datas = [d for d in upd_datas if d[3] in ["zc460", "zc429"] and d[1] == main_id]

        # -------------------------
        # Step 2：初始化空 DataFrame
        # -------------------------
        self.df_upd = pd.DataFrame({
            'ID': pd.Series(dtype='str'),
            'type': pd.Series(dtype='str'),
            'message_id': pd.Series(dtype='str'),
            'event time': pd.Series(dtype='str'),
            'direction': pd.Series(dtype='str'),
            'TrackingNumber': pd.Series(dtype='str')  # <<< 新增列
        })

        # -------------------------
        # Step 3：遍历数据并构建 DataFrame
        # -------------------------
        # Preconvert list-of-tuples into separate lists (fast)
        ids = []
        types_list = []
        msg_ids = []
        replay_times = []
        directions = []
        tracking_numbers = []

        # 遍历列表中的元组
        self.detail_upd_data = []

        # for data in upd_datas:
        #     # 提取元组中的所需元素
        #     response_data_id = data[1]
        #     type_value = data[3]  # 元组中的第3个元素
        #     replay_time = data[5]  # 元组中的第5个元素
        #     direction = data[6]
        #     message_id = data[10]
        #
        #     # 创建一个新的DataFrame用于存储当前行
        #     new_row = pd.DataFrame({
        #         'ID': [response_data_id],
        #         'type': [type_value],
        #         'message_id': [message_id],
        #         'event time': [replay_time],
        #         'direction': [direction]
        #     })
        #
        #     # 使用pd.concat来合并DataFrame
        #     self.df_upd = pd.concat([self.df_upd, new_row], ignore_index=True)
        #
        #     # 将第1和第4个元素以列表的形式存入self.detail_response_data
        #     self.detail_upd_data.append(
        #         [data[0], data[4], data[1], data[10], data[6], data[2], data[3]])  # data[0]是第1个元素，data[3]是第4个元素

        zc460_main_ids = []
        zc460_sub_ids = []
        zc429_main_ids = []

        for row in upd_datas:
            if row[3] == "zc460":  # type
                zc460_main_ids.append(row[1])
                zc460_sub_ids.append(row[2])
            elif row[3] == "zc429":
                zc429_main_ids.append(row[1])

        airwaybill_map_460 = db.get_airwaybill_by_main_ids(self.username, zc460_main_ids)
        tracking_number_map = db.get_tracking_numbers_by_sub_ids(self.username, zc460_sub_ids)
        airwaybill_map_429 = db.get_airwaybill_by_main_ids(self.username, zc429_main_ids)

        for row in upd_datas:
            _id = row[1]
            _type = row[3]
            _msgid = row[10]
            _replay_time = row[5]
            _direction = row[6]

            if _type == "zc460":
                main_table_id = row[1]
                sub_table_id = row[2]
                # 从 SubeExcelTable 查 tracking_number（sequence）
                airwaybill = airwaybill_map_460.get(main_table_id, main_table_id)
                tracking_number = tracking_number_map.get(sub_table_id, sub_table_id)
                _id = airwaybill
                # 查询失败不要 None，显示空白
                tracking_numbers.append(tracking_number)
            elif _type == "zc429":
                main_table_id = row[1]
                airwaybill = airwaybill_map_429.get(main_table_id, main_table_id)
                _id = airwaybill
                tracking_numbers.append("")  # zc429 没有 tracking_number
            else:
                tracking_numbers.append("")  # 其他类型直接空白

            # Detail structure preserved
            self.detail_upd_data.append([
                row[0],
                row[4],
                row[1],
                row[10],
                row[6],
                row[2],
                row[3]
            ])

            ids.append(_id)
            types_list.append(_type)
            msg_ids.append(_msgid)
            replay_times.append(_replay_time)
            directions.append(_direction)

        self.df_upd = pd.DataFrame({
            "ID": ids,
            "type": types_list,
            "message_id": msg_ids,
            "event time": replay_times,
            "direction": directions,
            "TrackingNumber": tracking_numbers
        })

        # -------------------------
        # Step 4：构建 QTableWidget
        # -------------------------
        self.upd_table.setRowCount(len(self.df_upd))
        self.upd_table.setColumnCount(len(self.df_upd.columns))
        self.upd_table.setHorizontalHeaderLabels(self.df_upd.columns)

        # -------------------------
        # Step 5：填充 QTableWidget 数据
        # -------------------------
        self.upd_table.setUpdatesEnabled(False)
        self.upd_table.blockSignals(True)
        self.upd_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.upd_table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)

        values = self.df_upd.values
        rows, cols = values.shape

        for row in range(rows):
            for col in range(cols):
                # item_text = str(self.df_upd.iloc[row, col]) if pd.notna(self.df_upd.iloc[row, col]) else ''
                # item = QTableWidgetItem(item_text)
                val = "" if pd.isna(values[row][col]) else str(values[row][col])
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # 设置为只读
                self.upd_table.setItem(row, col, item)

        # 恢复 UI 更新
        self.upd_table.blockSignals(False)
        self.upd_table.setUpdatesEnabled(True)
        self.upd_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.upd_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

    def show_main_table(self):
        self.flag = 0

        for i in reversed(range(self.info_layout.count())):
            widget = self.info_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()  # 删除控件
                self.info_layout.removeWidget(widget)

        self.df_response = pd.DataFrame({
            'ID': pd.Series(dtype='str'),
            'type': pd.Series(dtype='str'),
            'message_id': pd.Series(dtype='str'),
            'event time': pd.Series(dtype='str'),
            'direction': pd.Series(dtype='str')
        })
        self.update_response_table()

        self.index_table = 0
        self.main_table.clearSelection()  # 清空选择
        self.toolbar_return_action.setEnabled(False)
        # self.sub_table.hide()
        self.sub_table_container.hide()
        self.main_table.show()

    def show_sub_table(self, row, target_sub_id = None):
        self.flag = 1
        self.df_response = pd.DataFrame({
            'ID': pd.Series(dtype='str'),
            'type': pd.Series(dtype='str'),
            'message_id': pd.Series(dtype='str'),
            'event time': pd.Series(dtype='str'),
            'direction': pd.Series(dtype='str')
        })
        self.update_response_table()

        self.df_sub = pd.DataFrame({
            'sub_id': pd.Series(dtype='str'),
            'IOSS': pd.Series(dtype='str'),
            'TrackingNumber': pd.Series(dtype='str'),
            'lrn': pd.Series(dtype='str'),
            'state': pd.Series(dtype='str'),
            'event time': pd.Series(dtype='str')
        })

        self.index_table = 1
        self.sub_table.clearSelection()  # 清空选择

        self.main_table_id = self.df_main_data_list[row][1]
        self.input_information = json.loads(self.df_main_data_list[row][4])

        self.df_sub_data_list = db.get_sub_tables_by_main_id(self.main_table_id, self.username)

        for data in self.df_sub_data_list:
            state = data[3]
            replay_time = data[4]
            IOSS = data[7]
            TrackingNumber = data[8]
            lrn = data[9]
            sub_id = data[2]

            # 追加新行到 df_main
            self.df_sub = pd.concat([self.df_sub, pd.DataFrame({
                'sub_id': [sub_id],
                'IOSS': [IOSS],
                'TrackingNumber': [TrackingNumber],
                'lrn': [lrn],
                'state': [state],
                'event time': [replay_time]
            })], ignore_index=True)

        # print(self.df_sub)

        self.update_sub_table()

        self.main_table.hide()
        # self.sub_table.show()
        self.sub_table_container.show()
        self.toolbar_return_action.setEnabled(True)

        if target_sub_id is not None:
            for i, data in enumerate(self.df_sub_data_list):
                # 这里 data[0] 是 sub_id
                if str(data[2]) == str(target_sub_id):
                    self.sub_table.selectRow(i)
                    self.sub_table.scrollToItem(self.sub_table.item(i, 0))
                    break

    def create_toolbar(self):
        toolbar = QToolBar("ToolBar", self)
        self.addToolBar(toolbar)

        # 创建现有工具栏按钮
        self.toolbar_new_action = QAction('New', self)
        self.toolbar_return_action = QAction('Return', self)

        # 将现有按钮添加到工具栏中
        toolbar.addAction(self.toolbar_new_action)
        toolbar.addSeparator()  # 添加分隔符
        toolbar.addAction(self.toolbar_return_action)
        # exit_action = QAction('Exit', self)
        # exit_action.triggered.connect(self.close)
        # toolbar.addAction(exit_action)

        # 创建新按钮
        self.send_zc415_action = QAction('Send ZC415', self)
        self.send_reply_action = QAction('Send Reply Message', self)
        self.generate_pdf_and_xml_action = QAction('Generate PDF/XML', self)
        self.refresh_action = QAction('Refresh', self)

        self.change_cl380_action = QAction('Change CL380', self)

        self.clear_data_action = QAction('Filter Data', self)
        self.delete_timed_out_data_action = QAction('Delete Timed Out Data', self)

        self.import_dictionary_action = QAction('Import Dictionary', self)

        self.exit_login = QAction('Exit', self)

        # 添加新按钮到工具栏
        toolbar.addSeparator()  # 再次添加分隔符以将新按钮与现有按钮分开
        toolbar.addAction(self.send_zc415_action)
        toolbar.addAction(self.send_reply_action)
        toolbar.addAction(self.generate_pdf_and_xml_action)
        toolbar.addAction(self.refresh_action)
        toolbar.addSeparator()  # 再次添加分隔符以将新按钮与现有按钮分开
        toolbar.addAction(self.change_cl380_action)
        toolbar.addSeparator()  # 再次添加分隔符以将新按钮与现有按钮分开
        toolbar.addAction(self.clear_data_action)
        toolbar.addAction(self.delete_timed_out_data_action)

        # 获取QToolButton并设置样式
        self.tool_button = toolbar.widgetForAction(self.clear_data_action)
        self.tool_button = toolbar.widgetForAction(self.delete_timed_out_data_action)

        toolbar.addSeparator()  # 再次添加分隔符以将新按钮与现有按钮分开
        toolbar.addAction(self.import_dictionary_action)
        toolbar.addSeparator()  # 再次添加分隔符以将新按钮与现有按钮分开
        toolbar.addAction(self.exit_login)

        # 连接新按钮的触发信号
        self.send_zc415_action.triggered.connect(self.send_zc415)
        self.send_reply_action.triggered.connect(self.send_reply_message)
        self.generate_pdf_and_xml_action.triggered.connect(self.generate_pdf_and_xml)
        self.refresh_action.triggered.connect(self.refresh_data)

        self.change_cl380_action.triggered.connect(self.change_cl380)

        self.clear_data_action.triggered.connect(self.clear_data)
        self.delete_timed_out_data_action.triggered.connect(self.delete_timed_out_data)

        self.import_dictionary_action.triggered.connect(self.import_dictionary)

        self.exit_login.triggered.connect(self.exit_user_login)

        # 连接现有按钮的触发信号
        self.toolbar_new_action.triggered.connect(self.create_new)
        self.toolbar_return_action.triggered.connect(self.show_main_table)

        # 初始时禁用编辑和删除按钮
        self.toolbar_return_action.setEnabled(False)
        self.delete_timed_out_data_action.setEnabled(False)

    def send_zc415(self):
        # 创建并显示发送对话框
        dialog = SendZC415(self.username, self.token)
        dialog.exec_()
        self.show_main_table()
        self.update_main_table()
        self.update_upd_table()

    def send_reply_message(self):
        dialog = SendMessageByMainID(self.username, self.token)
        selection_opt = None
        if dialog.exec_() == QDialog.Accepted:
            selection_opt = dialog.get_selection_opt()
        self.show_main_table()
        self.update_main_table()
        if selection_opt == 'upd':
            QMessageBox.warning(self, "Warning", f'Please manually refresh upd!')
            self.refresh_data()

    def generate_pdf_and_xml(self):
        dialog = GeneratePDFAndXML(self.username, self.token)
        if dialog.exec_() == QDialog.Accepted:
            print("对话框成功关闭")
        else:
            print("对话框被取消")
        pass

    def refresh_data(self):
        dialog = RefreshByMainId(self.username, self.token)
        if dialog.exec_() == QDialog.Accepted:
            self.show_main_table()
            self.update_main_table()
            self.update_upd_table()
            print("对话框成功关闭")
        else:
            print("对话框被取消")
        pass

    def change_cl380(self):
        dialog = ChangeCL380ByMainID(self.username, self.token)
        dialog.exec_()

    # def clear_data(self):
    #     dialog = ClearData(self.username, self.token)
    #     dialog.exec_()
    #     self.update_upd_table()
    def clear_data(self):
        dialog = DateTimeDialog(self.time_flag)
        if dialog.exec_() == QDialog.Accepted:
            self.datetime = dialog.get_datetime()
            self.time_flag = dialog.get_time_flag()
            print(self.datetime)
            print(self.time_flag)
            self.update_upd_table()

    def delete_timed_out_data(self):
        reply = QMessageBox.question(self, 'Confirm Delete',
                                     'Are you sure you have backed up data older than 30 days? Click Yes to delete the relevant overdue data.',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            data = {
                'username': self.username,
                'main_ids': self.list_time_out_ids
            }
            response_status_code = http_client.delete_corresponding_data_by_main_ids(self.token, data)

            if response_status_code == 200:
                if db.delete_data_from_related_tables(self.list_time_out_ids):
                    if self.tool_button:
                        self.tool_button.setStyleSheet("QToolButton {}")
                        self.tool_button.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)  # 设置文字位置，视需要而定
                    self.delete_timed_out_data_action.setEnabled(False)
                    QMessageBox.warning(self, "Success", f'Deleting related data successfully!')
                else:
                    QMessageBox.warning(self, "Error", f'Failed to delete related data locally, please try again!')
            else:
                QMessageBox.warning(self, "Error", f'The server failed to delete related data, please try again!')

        self.show_main_table()
        self.update_main_table()
        self.update_upd_table()

    def import_dictionary(self):
        dialog = DictionarySelectorDialog()
        dialog.exec_()

    def exit_user_login(self):
        self.username = None
        self.token = None

        QMessageBox.information(self, "Loginout Successful", "You have successfully logged out!")

        self.login()

        self.update_main_table()

    def create_new(self):
        # new_sads = []
        create_sads_dialog = CreateSADs(self.username, self)
        # i = 1
        if create_sads_dialog.exec_() == QDialog.Accepted:
            list_new_keys = create_sads_dialog.get_data_keys()
            input_information = create_sads_dialog.get_input_information()
            selected_files_path = create_sads_dialog.get_selected_files()

            print(selected_files_path)

            key_map = {}

            system_to_user = dict(zip(self.list_fixed_key, list_new_keys))
            user_norm_to_system = {normalize(user_col): system_col
                                   for system_col, user_col in system_to_user.items()}

            for file_path in selected_files_path:
                try:
                    input_file = pd.read_excel(file_path, dtype={'TrackingNumber': str, 'ConsigneeNameID': str,
                                                                 'ConsigneeName': str, 'HSCode': str})

                    # === 列名标准化处理 ===
                    # normalized_columns = {}
                    # for col in input_file.columns:
                    #     col_lower = col.lower()
                    #     if col_lower == "dsk":
                    #         normalized_columns[col] = "Dsk"
                    #     elif col_lower in ["invoice number", "invoicenumber"]:
                    #         normalized_columns[col] = "InvoiceNumber"
                    # input_file.rename(columns=normalized_columns, inplace=True)

                    for col in input_file.columns:
                        col_norm = normalize(col)
                        if col_norm in user_norm_to_system:
                            key_map[col] = user_norm_to_system[col_norm]  # Excel列名 -> 系统列名

                    input_file.rename(columns=key_map, inplace=True)

                    # 将 nan 值替换为空字符串
                    input_file.fillna('', inplace=True)
                    input_file = input_file.applymap(lambda x: x.strip() if isinstance(x, str) else x)

                    # === 第一步：对 input_file 扫描 TrackingNumber 的首次出现行号 ===
                    tracking_index_map = {}
                    next_idx = 1
                    for _, row in input_file.iterrows():
                        tn = str(row.get("TrackingNumber", "")).strip()
                        if tn and tn not in tracking_index_map:
                            tracking_index_map[tn] = next_idx
                            next_idx += 1

                    # Check goodsitem previous document
                    for section in ["goodsitem previous document", "goodshipment previous document"]:
                        if input_information.get(section):
                            for item in input_information[section]:
                                if str(item.get("reference number", "")).strip() == "1":
                                    if "Dsk" not in input_file.columns:
                                        self.show_error(
                                            f"{section} default reading rule requires the Dsk column. "
                                            "Please remove related content if missing.")
                                        return False
                                    if input_file["Dsk"].replace('', pd.NA).isna().any():
                                        self.show_error(
                                            f"{section} previous document default reading rule requires the Dsk column, "
                                            f"but it is empty. Please remove related content.")
                                        return False

                    # Check goodsitem supporting document & goodshipment supporting document
                    for section in ["goodsitem supporting document", "goodshipment supporting document"]:
                        if input_information.get(section):
                            for item in input_information[section]:
                                if str(item.get("reference number", "")).strip() == "1":
                                    if "InvoiceNumber" not in input_file.columns:
                                        self.show_error(
                                            f"{section} default reading rule requires the InvoiceNumber column. "
                                            f"Please remove related content if missing.")
                                        return False
                                    if input_file["InvoiceNumber"].replace('', pd.NA).isna().any():
                                        self.show_error(
                                            f"{section} default reading rule requires the InvoiceNumber column, "
                                            f"but it is empty. Please remove related content.")
                                        return False

                    # Check goodshipment/goodsitem transport document reference=1
                    for section in ["goodshipment transport document", "goodsitem transport document"]:
                        if input_information.get(section):
                            for item in input_information[section]:
                                if str(item.get("reference number", "")).strip() == "1":
                                    if "AirWayBill" not in input_file.columns:
                                        self.show_error(
                                            f"{section} default reading rule requires the AirWayBill column. Please "
                                            f"remove related content if missing.")
                                        return False
                                    if input_file["AirWayBill"].replace('', pd.NA).isna().any():
                                        self.show_error(
                                            f"{section} default reading rule requires the AirWayBill column, but it "
                                            f"is empty. Please remove related content.")
                                        return False

                    # Check goodshipment/goodsitem transport document reference=2
                    for section in ["goodshipment transport document", "goodsitem transport document"]:
                        if input_information.get(section):
                            for item in input_information[section]:
                                if str(item.get("reference number", "")).strip() == "2":
                                    if "Box Number" not in input_file.columns:
                                        self.show_error(
                                            f"{section} default reading rule requires the Box Number column. Please "
                                            f"remove related content if missing.")
                                        return False
                                    if input_file["Box Number"].replace('', pd.NA).isna().any():
                                        self.show_error(
                                            f"{section} default reading rule requires the Box Number column, but it "
                                            f"is empty. Please remove related content.")
                                        return False
                except Exception as e:
                    self.show_error("No such file.")
                    return False

                try:
                    input_file_backup = input_file.copy(deep=True)
                    input_file = manifest.align_hscode(input_file)
                except manifest.HscodeError as e:
                    self.show_error(e.__str__())
                    return False

                data_negative = manifest.check_hscode(input_file)
                if not data_negative.empty:
                    # 提取目录和文件名
                    directory = os.path.dirname(file_path)
                    file_name = os.path.basename(file_path)
                    base_name, ext = os.path.splitext(file_name)

                    # 构造新文件名，用于保存标红的输入数据
                    new_file_name = f"{base_name}_Fixing_Errors{ext}"
                    new_file_path = os.path.join(directory, new_file_name)

                    # 创建Excel工作簿和工作表
                    wb = Workbook()
                    ws = wb.active

                    # 写入列名作为表头
                    for c_idx, col_name in enumerate(input_file.columns, 1):
                        cell = ws.cell(row=1, column=c_idx, value=col_name)

                    # 将data_negative转换为一行行的列表，方便匹配
                    negative_rows = data_negative.values.tolist()

                    # 遍历input_file，将数据写入工作表
                    for r_idx, row in enumerate(input_file.values, 2):  # 从第二行开始写数据
                        # 如果当前行在data_negative中，标记为红色
                        if row.tolist() in negative_rows:
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                # 设置字体颜色为红色
                                cell.font = Font(color="FF0000")
                        else:
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # 保存带标红的文件
                    wb.save(new_file_path)

                    # 额外保存data_negative为单独的错误文件
                    errors_file_name = f"{base_name}_Errors{ext}"
                    errors_file_path = os.path.join(directory, errors_file_name)

                    # 将data_negative保存为新的Excel文件
                    data_negative.to_excel(errors_file_path, index=False)

                    # 显示错误提示和相关信息
                    self.show_error(f'{data_negative.shape[0]} items with negative HScode detected.')
                    self.show_negative_goods_items(data_negative)

                    return False

                files = manifest.process_manifests(input_file_backup, file_path)

                processed_rows = []

                for file in files:
                    # key_map = {}
                    # for col in file.columns:
                    #     col_norm = normalize(col)
                    #     if col_norm in user_norm_to_system:
                    #         key_map[col] = user_norm_to_system[col_norm]  # Excel列名 -> 系统列名
                    #
                    # file.rename(columns=key_map, inplace=True)

                    # 选取要保留的系统列（不存在的列直接舍弃）
                    useful_cols = [c for c in self.list_fixed_key if c in file.columns]

                    # 4. 添加 TrackingNumberIndex（如果 TrackingNumber 也被舍弃，则需要判断）
                    if "TrackingNumber" in file.columns:
                        file["TrackingNumberIndex"] = (
                            file["TrackingNumber"].astype(str).str.strip().map(tracking_index_map)
                        ).fillna(-1).astype(int)
                        useful_cols.append("TrackingNumberIndex")

                    # 5. 转 dict
                    processed_rows.append(file[useful_cols].to_dict(orient="records"))

                sub_tables_additional_data = [input_information['goodsitem previous document'],
                                              input_information['goodsitem additional information'],
                                              input_information['goodsitem supporting document'],
                                              input_information['goodsitem additional reference'],
                                              input_information['goodsitem transport document']]
                db.store_excel_data(input_information, processed_rows, sub_tables_additional_data, self.username)
            self.update_main_table()
            self.show_main_table()
        else:
            print("Dialog rejected")

    def show_details(self, index):
        row = index.row()
        self.sub_table_id = self.df_sub_data_list[row][2]

        df_sub_data_list = db.get_sub_table_data(self.sub_table_id, self.username)
        data_rows = [json.loads(data[2]) for data in df_sub_data_list]

        manual_datas = [[json.loads(data[i]) for i in range(3, 8)] for data in df_sub_data_list]

        main_id_for_sub_table_data = [data[1] for data in df_sub_data_list]

        # print('######')
        # print(len(self.df_sub_data_list))
        # print(self.data_rows)
        # data_rows = self.list_data_rows[data['sequence number'] - 1]
        # print(data_rows)
        if len(data_rows) == 0:
            QMessageBox.warning(self, "Error", "No data.")
            return  # 必须加 return 确保初始化不会继续

        detail_dialog = SADWindow(self, data_rows=data_rows, input_information=self.input_information,
                                  manual_datas=manual_datas, main_table_id=self.main_table_id,
                                  sub_table_id=self.sub_table_id, main_id_for_sub_table_data=main_id_for_sub_table_data,
                                  username=self.username)
        if detail_dialog.exec_() == QDialog.Accepted:
            return_list = detail_dialog.get_some_return_information()
            self.update_main_table()
            # for t in return_list:
            #     print(t)

    def show_maximized(self):
        self.showMaximized()

    def show_error(self, msg):
        msg_box = QMessageBox(QMessageBox.Critical, "Error", msg)
        msg_box.exec_()

    def show_negative_goods_items(self, data_negative):
        # Create a new window
        self.table_window = QMainWindow()
        self.table_window.setWindowTitle("Negative Goods Items")
        self.table_window.setGeometry(200, 200, 800, 400)

        # Create a central widget and set layout
        central_widget = QWidget()
        self.table_window.setCentralWidget(central_widget)
        layout = QVBoxLayout()
        central_widget.setLayout(layout)

        # Create a table widget
        table_widget = QTableWidget()
        table_widget.setRowCount(data_negative.shape[0])
        table_widget.setColumnCount(3)
        table_widget.setHorizontalHeaderLabels(["GoodsItemNumber", "HSCode", "DescriptionGoods"])
        table_widget.horizontalHeader().setStretchLastSection(True)
        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # Populate the table widget with data
        for row, (_, item) in enumerate(data_negative.iterrows()):
            table_widget.setItem(row, 0, QTableWidgetItem(str(item.get('GoodsItemNumber'))))
            table_widget.setItem(row, 1, QTableWidgetItem(str(item.get('HSCode'))))
            table_widget.setItem(row, 2, QTableWidgetItem(str(item.get('DescriptionGoods'))))

        # Create a scroll area and set the table widget as its widget
        scroll_area = QScrollArea()
        scroll_area.setWidget(table_widget)
        scroll_area.setWidgetResizable(True)

        # Add the scroll area to the layout
        layout.addWidget(scroll_area)

        # Show the new window
        self.table_window.show()

    def login(self):
        login_dialog = LoginDialog()
        if login_dialog.exec_() == QDialog.Accepted:
            # 获取登录成功的用户名和 token
            self.username = login_dialog.username
            self.token = login_dialog.token
            self.exit_login.setText('Exit ' + self.username)
            if self.token:
                success, msg = http_client.synchronize_data(self.token, self.username)
                if success:
                    QMessageBox.information(self, "Success", msg)
                else:
                    QMessageBox.warning(self, "Failure", msg)

                self.list_time_out_ids = db.get_sequences_older_than_30_days(self.username)
                if len(self.list_time_out_ids['main_table_ids']) + len(self.list_time_out_ids['xml_table_ids']) > 0:
                    if self.tool_button:
                        self.tool_button.setStyleSheet("QToolButton {color: red; font-weight: bold;}")
                        self.tool_button.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)  # 设置文字位置，视需要而定
                    self.delete_timed_out_data_action.setEnabled(True)
            self.show_main_table()
            self.update_main_table()
            self.update_upd_table()
            print("登录成功")
        else:
            print("登录失败")
            sys.exit()  # 退出程序

    # def synchronize_user_information(self):
    #     pass


if __name__ == '__main__':
    app = QApplication(sys.argv)
    mainWin = MainWindow()
    mainWin.show()
    sys.exit(app.exec_())
